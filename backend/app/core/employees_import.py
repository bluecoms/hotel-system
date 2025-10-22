# app/core/employees_import.py
# -*- coding: utf-8 -*-
from __future__ import annotations
from typing import List, Tuple, Optional, Dict, Any
from fastapi import UploadFile, HTTPException
from pydantic import BaseModel
import csv, io, re
from datetime import date as _date

# ─────────────────────────────────────────────────────────────────────────────
# 텍스트/테이블 파서 (기존 유지)

def guess_text(raw: bytes) -> str:
    for enc in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return raw.decode(enc)
        except Exception:
            pass
    return raw.decode("utf-8", errors="ignore")

def html_table_to_rows(html_text: str) -> list[list[str]]:
    import html as _html
    tables = re.findall(r"<table.*?>.*?</table>", html_text, flags=re.I|re.S)
    if not tables: return []
    def table_to_rows(t: str) -> list[list[str]]:
        rows=[]
        for tr in re.findall(r"<tr.*?>(.*?)</tr>", t, flags=re.I|re.S):
            cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, flags=re.I|re.S)
            row=[]
            for c in cells:
                txt = re.sub(r"<.*?>","",c,flags=re.S)
                txt = _html.unescape(txt).strip()
                row.append(txt)
            if any(x.strip() for x in row): rows.append(row)
        return rows
    cand = [table_to_rows(t) for t in tables]
    cand.sort(key=lambda r: (len(r), max((len(x) for x in r), default=0)), reverse=True)
    return cand[0] if cand else []

def xlsx_to_rows(raw: bytes) -> list[list[str]]:
    try:
        import openpyxl, io as _io
    except Exception:
        raise HTTPException(415, "XLSX not supported. Install openpyxl or upload CSV.")
    wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out=[]
    for row in ws.iter_rows(values_only=True):
        out.append(["" if v is None else str(v) for v in row])
    return out

def xls_to_rows(raw: bytes) -> list[list[str]]:
    try:
        import xlrd  # 1.2.0 권장
    except Exception:
        raise HTTPException(415, "XLS not supported. Install xlrd==1.2.0")
    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    out = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            v = sh.cell_value(r, c)
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            row.append("" if v is None else str(v))
        out.append(row)
    return out

def to_rows(file: UploadFile) -> list[list[str]]:
    raw = file.file.read()

    # HTML <table>
    if b"<table" in raw[:4096].lower() and b"</table>" in raw[:4096].lower():
        txt = guess_text(raw)
        rows = html_table_to_rows(txt)
        if rows: return rows

    # XLSX
    if (file.filename or "").lower().endswith(".xlsx") or raw[:2] == b"PK":
        return xlsx_to_rows(raw)

    # XLS (OLE)
    if (file.filename or "").lower().endswith(".xls") or raw[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return xls_to_rows(raw)

    # CSV/TXT (기본)
    txt = guess_text(raw)
    return list(csv.reader(io.StringIO(txt)))

# ─────────────────────────────────────────────────────────────────────────────
# 정규화/마스킹 유틸 (기존 유지)

def norm_header(h: str) -> str:
    s = (h or "").strip().lower()
    s = s.replace(" ", "").replace("\u00a0","")
    return s

def mask_rrn(s: str) -> str:
    """801125-1234567 → 801125-1** 식으로 마스킹"""
    if not s: return ""
    s = re.sub(r"[^0-9\-*]", "", s)
    m = re.search(r"(\d{6})[\-]?\s*([1-4])", s)
    if not m: return ""
    return f"{m.group(1)}-{m.group(2)}**"

def mask_account(s: str) -> Tuple[str, str]:
    """계좌번호 마스킹 → (***-***-1234, 1234)"""
    if not s: return ("","")
    digits = re.sub(r"[^\d]", "", s)
    if not digits: return ("","")
    last4 = digits[-4:] if len(digits)>=4 else digits
    masked = f"{'*'*3}-{'*'*3}-{last4}"
    return (masked, last4)

def smart_date(s: str) -> Optional["_date"]:
    import datetime as dt
    ss = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%y-%m-%d", "%y.%m.%d", "%y/%m/%d"):
        try:
            return dt.datetime.strptime(ss, fmt).date()
        except:
            pass
    if ss.isdigit() and len(ss) == 8:  # 20240831
        try:
            return dt.datetime.strptime(ss, "%Y%m%d").date()
        except:
            pass
    return None

# ─────────────────────────────────────────────────────────────────────────────
# Pydantic 모델 (라우터에서 import)

class PreviewRow(BaseModel):
    name: str = ""
    department: str = ""
    position: str = ""
    hire_date: Optional[_date] = None
    rrn_masked: str = ""
    phone: str = ""
    bank_name: str = ""
    bank_account_masked: str = ""
    bank_last4: str = ""
    email: str = ""
    raw: Dict[str, Any] = {}

class ImportResult(BaseModel):
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: List[str] = []

# ─────────────────────────────────────────────────────────────────────────────
# 컬럼 매핑

_ALIASES: Dict[str, str] = {
    # 이름
    "성명": "name", "이름": "name", "name": "name",
    # 부서/직책
    "부서": "department", "department": "department", "팀": "department",
    "직책": "position", "직위": "position", "position": "position",
    # 입사일
    "입사일": "hire_date", "입사일자": "hire_date", "hiredate": "hire_date", "입사": "hire_date",
    # 주민/전화
    "주민번호": "rrn", "주민등록번호": "rrn", "rrn": "rrn",
    "전화": "phone", "연락처": "phone", "휴대폰": "phone", "mobile": "phone", "phone": "phone",
    # 계좌/은행
    "계좌": "account", "계좌번호": "account", "account": "account",
    "은행": "bank_name", "bank": "bank_name", "은행명": "bank_name",
    # 이메일
    "이메일": "email", "email": "email",
}

def _header_map(headers: List[str]) -> Dict[int, str]:
    m: Dict[int, str] = {}
    for i, h in enumerate(headers):
        key = norm_header(h)
        key = _ALIASES.get(key, key)  # alias 적용
        m[i] = key
    return m

# ─────────────────────────────────────────────────────────────────────────────
# 미리보기

def preview_employees_from_file(file: UploadFile) -> List[PreviewRow]:
    rows = to_rows(file)
    if not rows:
        raise HTTPException(422, "parsed-0-rows")
    headers = rows[0]
    body = rows[1:]

    hmap = _header_map(headers)
    previews: List[PreviewRow] = []

    for r in body:
        obj: Dict[str, Any] = {}
        for idx, key in hmap.items():
            if idx < len(r):
                obj[key] = (r[idx] or "").strip()

        pr = PreviewRow(
            name=obj.get("name",""),
            department=obj.get("department",""),
            position=obj.get("position",""),
            hire_date=smart_date(obj.get("hire_date","")),
            rrn_masked=mask_rrn(obj.get("rrn","")),
            phone=obj.get("phone",""),
            bank_name=obj.get("bank_name",""),
            bank_account_masked=mask_account(obj.get("account",""))[0],
            bank_last4=mask_account(obj.get("account",""))[1],
            email=obj.get("email",""),
            raw=obj,
        )
        # 완전 빈 행은 스킵
        if any([pr.name, pr.department, pr.position, pr.phone, pr.email, pr.bank_last4]):
            previews.append(pr)

    return previews

# ─────────────────────────────────────────────────────────────────────────────
# 가져오기 (DB upsert: email > phone > name 기준)

def import_employees_from_csv(file: UploadFile, db) -> ImportResult:
    try:
        # 미리보기 기반으로 표준화 먼저
        previews = preview_employees_from_file(file)
    except HTTPException as e:
        raise
    except Exception as e:
        raise HTTPException(400, f"preview-failed: {e}")

    # 모델 임포트 (존재 가정)
    try:
        from app.models.employee import Employee  # type: ignore
    except Exception:
        # 모델이 없으면 미리보기만 통과하게 하고, DB는 건드리지 않음
        return ImportResult(created=0, updated=0, skipped=len(previews), errors=["Employee model not found"])

    created = 0
    updated = 0
    skipped = 0
    errors: List[str] = []

    for pr in previews:
        try:
            # 우선 키 결정
            q = None
            if pr.email:
                q = db.query(Employee).filter(Employee.email == pr.email).first()
            if q is None and pr.phone:
                q = db.query(Employee).filter(Employee.phone == pr.phone).first()
            if q is None and pr.name:
                q = db.query(Employee).filter(Employee.name == pr.name).first()

            if q is None:
                # 생성
                obj = Employee()
                # 속성 있으면 세팅 (필드 유연성)
                if hasattr(obj, "name"): obj.name = pr.name
                if hasattr(obj, "department"): obj.department = pr.department
                if hasattr(obj, "position"): obj.position = pr.position
                if hasattr(obj, "hire_date"): obj.hire_date = pr.hire_date
                if hasattr(obj, "phone"): obj.phone = pr.phone
                if hasattr(obj, "email"): obj.email = pr.email
                if hasattr(obj, "bank_name"): obj.bank_name = pr.bank_name
                if hasattr(obj, "bank_last4"): obj.bank_last4 = pr.bank_last4
                # 필요 시 원본계좌 전체 저장 필드가 있으면 pr.raw.get("account")
                db.add(obj)
                created += 1
            else:
                # 업데이트 (필드 존재 시에만)
                changed = False
                if hasattr(q, "department") and q.department != pr.department:
                    q.department = pr.department; changed = True
                if hasattr(q, "position") and q.position != pr.position:
                    q.position = pr.position; changed = True
                if hasattr(q, "hire_date") and q.hire_date != pr.hire_date:
                    q.hire_date = pr.hire_date; changed = True
                if hasattr(q, "phone") and pr.phone and q.phone != pr.phone:
                    q.phone = pr.phone; changed = True
                if hasattr(q, "email") and pr.email and q.email != pr.email:
                    q.email = pr.email; changed = True
                if hasattr(q, "bank_name") and q.bank_name != pr.bank_name:
                    q.bank_name = pr.bank_name; changed = True
                if hasattr(q, "bank_last4") and q.bank_last4 != pr.bank_last4:
                    q.bank_last4 = pr.bank_last4; changed = True
                if changed:
                    updated += 1
                else:
                    skipped += 1
        except Exception as e:
            errors.append(str(e))
            skipped += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"commit-failed: {e}")

    return ImportResult(created=created, updated=updated, skipped=skipped, errors=errors)
