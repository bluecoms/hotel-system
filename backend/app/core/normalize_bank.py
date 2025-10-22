# -*- coding: utf-8 -*-
# ============================================================================
# File      : app/core/normalize_bank.py
# Version   : 2025.10-30 · v1.2 (Keep-Logic · Commented · Safer)
# Purpose   : Hotel Admin — Banking Upload Normalizer (입출금 원본 전처리)
# ----------------------------------------------------------------------------
# 목적:
#   • 업로드된 원본 파일(.csv/.xls/.xlsx/HTML table)을 "정규 CSV 텍스트"로 변환
#   • 다양한 은행 포맷을 공통 Canon 헤더로 매핑 (date,time,direction,...)
#   • SSOT 병합 엔진 전 단계(업로드 처리)에서 사용하는 독립 유틸리티
# ----------------------------------------------------------------------------
# 변경 로그:
#   v1.2 (2025-10-30)
#     ✅ 주석 대폭 보강 (역할/주의점/에러 처리)
#     ✅ XLS/XLSX/HTML 감지 보완 및 예외 메시지 정리
#     ✅ 빈 데이터(0 rows) 업로드 시 422 오류 반환
#     ✅ amount/날짜/방향 정규화 방어 로직 소폭 보강
#   v1.1
#     - key 후보 확장(_pick), 괄호/부호 금액 처리 보강
#   v1.0
#     - 최초 도입
# ============================================================================
from __future__ import annotations

import io
import csv
import datetime as dt
from typing import List, Tuple, Optional, Dict, Any

from fastapi import HTTPException

# 파일 매직 (확장자 대신 바이너리 시그니처로 판별)
OLE_MAGIC = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"  # .xls (OLE)
ZIP_MAGIC = b"PK"                                 # .xlsx (zip)

# ============================================================================
# 공통 유틸
# ----------------------------------------------------------------------------
def _bytes_to_text_guess(raw: bytes) -> str:
    """
    업로드 바이트 → 텍스트 (인코딩 추정)
    - 가장 흔한 UTF-8 / CP949 / EUC-KR 순으로 시도
    - 모두 실패 시 UTF-8 ignore
    """
    for enc in ("utf-8-sig", "cp949", "euc-kr", "ms949"):
        try:
            return raw.decode(enc)
        except Exception:
            pass
    return raw.decode("utf-8", errors="ignore")


def _looks_like_xlsx(filename: Optional[str], head: bytes) -> bool:
    """파일명/시그니처 기반 .xlsx 판별 (zip)"""
    fn = (filename or "").lower()
    return fn.endswith(".xlsx") or head.startswith(ZIP_MAGIC)


def _looks_like_xls(head: bytes) -> bool:
    """.xls (OLE) 여부"""
    return head.startswith(OLE_MAGIC)


def _xlsx_to_csv_text(raw: bytes) -> str:
    """
    XLSX(첫 시트) → CSV 텍스트
    - openpyxl 필요 (read_only=True 권장)
    """
    try:
        import openpyxl, io as _io, csv as _csv  # type: ignore
    except Exception:
        raise HTTPException(415, "XLSX not supported. Install openpyxl.")

    wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = _io.StringIO()
    w = _csv.writer(out, lineterminator="\n")
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if v is None else str(v) for v in row])
    return out.getvalue()


def _xls_to_csv_text(raw: bytes) -> str:
    """
    XLS(첫 시트) → CSV 텍스트
    - xlrd 1.2.0 필요 (2.x는 xlsx 미지원)
    """
    try:
        import xlrd  # type: ignore  # 1.2.0
    except Exception:
        raise HTTPException(415, "XLS not supported. Install xlrd==1.2.0.")

    book = xlrd.open_workbook(file_contents=raw)
    sh = book.sheet_by_index(0)
    import io as _io, csv as _csv
    out = _io.StringIO()
    w = _csv.writer(out, lineterminator="\n")
    for r in range(sh.nrows):
        vals = []
        for c in range(sh.ncols):
            v = sh.cell_value(r, c)
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            vals.append("" if v is None else str(v))
        w.writerow(vals)
    return out.getvalue()


def _promote_header(csv_text: str) -> str:
    """
    CSV 헤더 추정 후 최상단으로 승격
    - 초반 60줄 내에서 '문자열 열 개수'가 많은 라인을 헤더로 간주
    - 완벽한 보장은 아님: 최대한 안전하게 시도
    """
    rows = csv_text.splitlines()
    if not rows:
        return csv_text
    best_i, best_score = 0, -10**9
    # 빈 줄/공백 라인 제거 고려
    for i, line in enumerate(rows[:60]):
        if not line.strip():
            continue
        parts = [c.strip() for c in line.split(",")]
        nonnum = sum(1 for c in parts if not c.replace(".", "", 1).isdigit())
        score = len(parts) + nonnum * 2
        if score > best_score:
            best_i, best_score = i, score
    out = "\n".join([rows[best_i]] + rows[best_i + 1 :])
    return out if out.endswith("\n") else (out + "\n")


def norm_amount(x: str) -> int:
    """
    통화 문자열 → 정수 원 단위
    - 콤마/원화기호/여러 공백 제거
    - (1,234) → -1234 / 1,234- → -1234 / '−,–' → '-'
    - 변환 실패 시 0
    """
    s = (x or "")
    s = (
        s.replace(",", "")
        .replace("₩", "")
        .replace("￦", "")
        .replace("원", "")
        .replace("\u00a0", "")
        .replace("\u2009", "")
        .replace(" ", "")
    )
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace("−", "-").replace("–", "-")
    if s.endswith("-") and s[:-1].replace(".", "", 1).isdigit():
        s = "-" + s[:-1]
    try:
        return int(round(float(s or "0")))
    except Exception:
        return 0


def _is_summary_row(r: dict) -> bool:
    """합계/잔액/소계 등 요약 라인 판별 → 본문에서 제외"""
    try:
        text = " ".join([str(v or "") for v in r.values()]).lower()
    except Exception:
        return False
    if not any(str(v or "").strip() for v in r.values()):
        return True
    return any(k in text for k in ["합계", "총계", "잔액", "이월", "소계", "총합", "누계", "계)"])


def _pick(r: dict, cand: List[str]) -> str:
    """여러 헤더 후보 중 존재/비어있지 않은 첫 값을 선택"""
    for k in cand:
        if k in r and r[k] is not None:
            v = str(r[k]).strip()
            if v != "":
                return v
    return ""


# ============================================================================
# 업로드 바이트 → CSV 텍스트
# ----------------------------------------------------------------------------
def read_bank_upload_to_csv_text(file) -> Tuple[str, str]:
    """
    업로드 파일(UploadFile) → CSV 텍스트 변환
    - HTML table 감지 시 원문 텍스트 반환(별도 HTML 파서는 쓰지 않음)
    - XLSX/XLS 각각 openpyxl/xlrd로 1시트만 덤프
    - 최종 CSV 텍스트와 '.csv' 확장자 반환
    """
    raw = file.file.read()
    head = raw[:8]
    low4k = raw[:4096].lower()

    # HTML table?
    if b"<table" in low4k and b"</table>" in low4k:
        txt = _bytes_to_text_guess(raw)
        csv_text = txt
    elif _looks_like_xlsx(file.filename, head):
        csv_text = _xlsx_to_csv_text(raw)
    elif _looks_like_xls(head):
        csv_text = _xls_to_csv_text(raw)
    else:
        csv_text = _bytes_to_text_guess(raw)

    csv_text = _promote_header(csv_text)
    if not csv_text.endswith("\n"):
        csv_text += "\n"
    return csv_text, ".csv"


# ============================================================================
# NH 등 통장 입/출금 공통 정규화
#  - 파일이 입금/출금 분리되어 있어도 컬럼 이름으로 auto 감지
#  - 반환: (정규 CSV 텍스트, preview_rows(최대 5줄), stats)
# ----------------------------------------------------------------------------
CANON_FIELDS = [
    "date",
    "time",
    "direction",
    "amount",
    "balance",
    "desc",
    "counterparty",
    "memo",
    "raw_ref",
]

def normalize_bank_csv(
    csv_text: str, default_direction: Optional[str] = None
) -> Tuple[str, List[str], Dict[str, Any]]:
    """
    은행 원본 CSV → Canon CSV
    - 헤더 후보를 스캔해 공통 필드로 매핑
    - direction/amount/balance 표준화
    - 결과 CSV 텍스트와 프리뷰(최대 5줄), 통계(rows/in/out) 반환
    - 유효 데이터 0건 시 422 오류 반환
    """
    rdr = csv.DictReader(io.StringIO(csv_text))
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=CANON_FIELDS, lineterminator="\n")
    w.writeheader()

    n_in, n_out, n_rows = 0, 0, 0
    preview: List[str] = []

    for r in rdr:
        # 요약/빈 행 skip
        if _is_summary_row(r):
            continue

        # 후보 헤더에서 값 선택
        date = _pick(r, ["거래일자", "거래일", "일자", "date", "DATE", "적요일자"]).replace(".", "-").replace("/", "-")
        time = _pick(r, ["거래시간", "시간", "time", "TIME", "시각"])
        desc = _pick(r, ["적요", "적요내용", "내용", "비고", "설명", "명세", "내역", "거래내용", "계좌적요"])
        cp = _pick(r, ["거래처", "받는분", "입금자", "출금처", "상대계좌", "상대명", "거래점", "지점"])
        memo = _pick(r, ["메모", "비고", "참고"])
        ref = _pick(r, ["거래번호", "전표번호", "문서번호", "거래일련번호", "번호"])

        amt_in = norm_amount(_pick(r, ["입금", "입금액", "입금금액", "DEPOSIT", "credit", "Credit", "+금액", "+"]))
        amt_out = norm_amount(_pick(r, ["출금", "출금액", "출금금액", "WITHDRAW", "debit", "Debit", "-금액", "-"]))
        bal = norm_amount(_pick(r, ["잔액", "BALANCE", "balance", "현재잔액", "잔고"]))

        # 방향/금액 결정
        direction = (default_direction or "").upper() or None
        amount = 0
        if amt_in and not amt_out:
            direction = "IN"
            amount = amt_in
            n_in += 1
        elif amt_out and not amt_in:
            direction = "OUT"
            amount = amt_out
            n_out += 1
        else:
            # 둘 다 0 또는 둘 다 존재 → default 사용, 없으면 금액이 0 이면 skip
            amount = amt_in or amt_out or 0
            if not direction:
                if amount == 0:
                    # 금액 정보가 전혀 없으면 유효 거래 아님
                    continue
                # 부호로 방향 추정 (괄호/부호 처리 후 norm_amount 에서 이미 반영됨)
                # 이 시점에서 amount > 0 이면 IN 으로 간주
                direction = "IN"

        # 날짜 파싱 보정 (yyyy-mm-dd)
        ds = (date or "").strip()
        if ds and len(ds) >= 8:
            for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%y-%m-%d", "%Y%m%d"):
                try:
                    ds = dt.datetime.strptime(date, fmt).strftime("%Y-%m-%d")
                    break
                except Exception:
                    continue

        row = {
            "date": ds or "",
            "time": time or "",
            "direction": (direction or "IN").upper()[:3],
            "amount": str(amount),
            "balance": str(bal or 0),
            "desc": desc or "",
            "counterparty": cp or "",
            "memo": memo or "",
            "raw_ref": ref or "",
        }
        w.writerow(row)
        n_rows += 1

        if len(preview) < 5:
            preview.append(",".join([row[k] for k in CANON_FIELDS]))

    # ✅ 유효 행이 하나도 없으면 422 반환 (사용자에게 명확한 피드백)
    if n_rows == 0:
        raise HTTPException(status_code=422, detail="유효한 거래 내역이 없습니다. (빈 파일 또는 인식 불가 형식)")

    return out.getvalue(), preview, {"rows": n_rows, "in": n_in, "out": n_out}
